import tkinter as tk
from tkinter import Label, Entry, Button, Frame, LabelFrame, messagebox
from PIL import Image, ImageTk
import os
from openpyxl import Workbook, load_workbook

# Function to clear previous results
def clear_previous_results():
    for widget in user_data_frame.grid_slaves():
        if int(widget.grid_info()["row"]) == 3:
            widget.grid_forget()

# Function to clear all input fields
def clear_entries():
    first_name_entry.delete(0, tk.END)
    last_name_entry.delete(0, tk.END)
    height_entry.delete(0, tk.END)
    weight_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    clear_previous_results()

# Function to save BMI data to an Excel file
def save_to_excel(first_name, last_name, height, weight, age, bmi=None, bmi_category=None):
    # Ask the user to choose the file location
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if file_path:  # Only proceed if the user selected a file path
        # Check if the file already exists
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Add headers if creating a new file
            sheet.append(["First Name", "Last Name", "Height (m)", "Weight (kg)", "Age", "BMI", "Category"])

        # Append the user data and BMI results
        sheet.append([first_name, last_name, height, weight, age, round(bmi, 2) if bmi else "", bmi_category if bmi_category else ""])
        workbook.save(file_path)

# Function to calculate BMI
def calculate_bmi():
    clear_previous_results()
    try:
        # Get user input and convert to appropriate types
        weight = float(weight_entry.get())
        height = float(height_entry.get())
        age = int(age_entry.get())
        first_name = first_name_entry.get()
        last_name = last_name_entry.get()

        # Validate input values
        if weight <= 0 or height <= 0:
            raise ValueError("Height and weight must be positive numbers.")

        # Calculate BMI
        bmi = weight / (height ** 2)

        # Classify BMI
        if bmi < 18.5:
            bmi_category = "Underweight"
        elif bmi < 24.9:
            bmi_category = "Normal weight"
        elif bmi < 29.9:
            bmi_category = "Overweight"
        else:
            bmi_category = "Obese"

        # Save to Excel
        save_to_excel(first_name, last_name, height, weight, age, bmi, bmi_category)

        # Display BMI result
        bmi_label = Label(user_data_frame, text=f"{first_name} {last_name}, your BMI is {bmi:.2f}, and you are {bmi_category}",
                          bg="#282c34", fg="#61dafb", font=("Helvetica", 14, "bold"))
        bmi_label.grid(row=3, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
    except ValueError as e:
        bmi_label = Label(user_data_frame, text=str(e), bg="#282c34", fg="red", font=("Helvetica", 14, "bold"))
        bmi_label.grid(row=3, column=0, columnspan=4, sticky="ew", padx=5, pady=5)

# Create the main window
window = tk.Tk()
window.title("BMI CALCULATOR")
window.configure(bg="#282c34")
window.geometry("500x400")

# Add an icon to the window using Pillow
image = Image.open("C:/Users/ZBOOK/Pictures/logo.png")
icon = ImageTk.PhotoImage(image)
window.iconphoto(False, icon)

# Configure the grid for the main window to be responsive
window.grid_rowconfigure(1, weight=1)
window.grid_columnconfigure(0, weight=1)

# Create a header label
header_label = Label(window, text="BMI Calculator", font=("Helvetica", 24, "bold"), bg="#61dafb", fg="#282c34")
header_label.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

# Create a frame within the main window
frame = Frame(window, bg="#282c34")
frame.grid(row=1, column=0, sticky="nsew")

# Configure the grid within the frame
frame.grid_rowconfigure(0, weight=1)
frame.grid_columnconfigure(0, weight=1)

# Create a labeled frame for user data
user_data_frame = LabelFrame(frame, text="User Data", bg="#61dafb", fg="#282c34", font=("Helvetica", 16, "bold"))
user_data_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

# Configure the grid within the user_data_frame to be responsive
for i in range(3):
    user_data_frame.grid_rowconfigure(i, weight=1)
for i in range(4):
    user_data_frame.grid_columnconfigure(i, weight=1)

# Create labels and entry fields for user data
first_name_label = Label(user_data_frame, text="First Name", bg="#61dafb", fg="#282c34", font=("Helvetica", 12))
first_name_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)

last_name_label = Label(user_data_frame, text="Last Name", bg="#61dafb", fg="#282c34", font=("Helvetica", 12))
last_name_label.grid(row=0, column=2, sticky="e", padx=5, pady=5)

first_name_entry = Entry(user_data_frame, font=("Helvetica", 12))
last_name_entry = Entry(user_data_frame, font=("Helvetica", 12))
first_name_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
last_name_entry.grid(row=0, column=3, sticky="ew", padx=5, pady=5)

height_label = Label(user_data_frame, text="Enter Height in Meters", bg="#61dafb", fg="#282c34", font=("Helvetica", 12))
height_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
weight_label = Label(user_data_frame, text="Enter Weight in Kilograms", bg="#61dafb", fg="#282c34", font=("Helvetica", 12))
weight_label.grid(row=1, column=2, sticky="e", padx=5, pady=5)

height_entry = Entry(user_data_frame, font=("Helvetica", 12))
height_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
weight_entry = Entry(user_data_frame, font=("Helvetica", 12))
weight_entry.grid(row=1, column=3, sticky="ew", padx=5, pady=5)

age_label = Label(user_data_frame, text="Enter your Age", bg="#61dafb", fg="#282c34", font=("Helvetica", 12))
age_label.grid(row=2, column=0, sticky="e", padx=5, pady=5)
age_entry = Entry(user_data_frame, font=("Helvetica", 12))
age_entry.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

# Create buttons to trigger BMI calculation and clear entries
calculate_button = Button(frame, text="Calculate", command=calculate_bmi, bg="#61dafb", fg="#282c34", font=("Helvetica", 14, "bold"))
calculate_button.grid(row=1, column=0, sticky="ew", padx=10, pady=10)

clear_button = Button(frame, text="Clear", command=clear_entries, bg="#61dafb", fg="#282c34", font=("Helvetica", 14, "bold"))
clear_button.grid(row=2, column=0, sticky="ew", padx=10, pady=10)

# Add hover effects to buttons
def on_enter(e):
    e.widget['background'] = '#282c34'
    e.widget['foreground'] = '#61dafb'

def on_leave(e):
    e.widget['background'] = '#61dafb'
    e.widget['foreground'] = '#282c34'

calculate_button.bind("<Enter>", on_enter)
calculate_button.bind("<Leave>", on_leave)

clear_button.bind("<Enter>", on_enter)
clear_button.bind("<Leave>", on_leave)

# Start the application
window.mainloop()
